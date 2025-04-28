import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def consultar_cpf(driver, cpf):
    wait = WebDriverWait(driver, 10)
    
    campo_pesquisa = wait.until(EC.presence_of_element_located((By.ID, 'cpfInput')))
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)

    botao_pesquisar = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-custom')]")))
    botao_pesquisar.click()

    status = wait.until(EC.presence_of_element_located((By.ID, 'statusLabel')))
    
    if status.text.strip().lower() == 'em dia':
        data_pagamento = driver.find_element(By.ID, 'paymentDate').text.split()[3]
        metodo_pagamento = driver.find_element(By.ID, 'paymentMethod').text.split()[3]
        return 'em dia', data_pagamento, metodo_pagamento
    else:
        return 'pendente', None, None

def main():
    
    planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
    pagina_clientes = planilha_clientes['Sheet1']
    
    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']
    
    driver = webdriver.Chrome()
    driver.get('https://consultcpf-devaprender.netlify.app/')

    for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
        nome, valor, cpf, vencimento = linha
        try:
            status, data_pagamento, metodo_pagamento = consultar_cpf(driver, cpf)

            if status == 'em dia':
                pagina_fechamento.append([nome, valor, cpf, vencimento, status, data_pagamento, metodo_pagamento])
            else:
                pagina_fechamento.append([nome, valor, cpf, vencimento, status])
        
        except Exception as e:
            print(f"Erro ao processar CPF {cpf}: {e}")
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'erro'])

    planilha_fechamento.save('planilha fechamento.xlsx')
    driver.quit()

if __name__ == "__main__":
    main()
        
        
         
